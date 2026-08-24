#!/usr/bin/env python3
"""
Export completed VSign UAT transactions into Verasys Excel format:
  Sr. No. | DATE | TXN ID

Only rows with signed PDF on disk are included (genuine completed txns).

Usage:
  python scripts/export-vsign-uat-excel.py
  python scripts/export-vsign-uat-excel.py --out "C:/Users/DELL/Desktop/50 eSign Transaction Format.xlsx"
  python scripts/export-vsign-uat-excel.py --limit 50 --merge "C:/Users/DELL/Desktop/50 eSign Transaction Format.xlsx"
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from pymongo import MongoClient
    from bson.objectid import ObjectId
except ImportError:
    print("Install: pip install pymongo", file=sys.stderr)
    sys.exit(1)


def load_env_mongo_uri(service_root: Path) -> str:
    env_path = service_root.parent.parent / ".env"
    uri = "mongodb://127.0.0.1:27017/draftnsign"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            if line.startswith("MONGO_URI="):
                uri = line.split("=", 1)[1].strip()
                break
    return uri


def fetch_completed_txns(mongo_uri: str, limit: int) -> list[dict]:
    client = MongoClient(mongo_uri)
    db = client.get_default_database()
    coll = db["signaturetransactions"]
    rows: list[dict] = []

    for doc in coll.find({}).sort("_id", -1):
        signed_path = doc.get("signedFilePath") or ""
        if not signed_path or not os.path.isfile(signed_path):
            continue
        oid = doc.get("_id")
        created = oid.generation_time if isinstance(oid, ObjectId) else datetime.now(timezone.utc)
        if created.tzinfo is None:
            created = created.replace(tzinfo=timezone.utc)
        local_dt = created.astimezone()
        rows.append(
            {
                "txn": str(doc.get("txn") or ""),
                "date": local_dt.replace(tzinfo=None),
            }
        )
        if len(rows) >= limit:
            break

    client.close()
    rows.sort(key=lambda r: r["txn"])
    return rows


def write_excel(out_path: Path, txns: list[dict], target_rows: int = 50) -> None:
    if out_path.exists():
        wb = load_workbook(out_path)
        ws = wb.active
    else:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(1, 1, "Sr. No.")
        ws.cell(1, 2, "DATE")
        ws.cell(1, 3, "TXN ID")

    for i in range(target_rows):
        row_idx = i + 2
        if i < len(txns):
            ws.cell(row_idx, 1, i + 1)
            ws.cell(row_idx, 2, txns[i]["date"])
            ws.cell(row_idx, 3, int(txns[i]["txn"]) if txns[i]["txn"].isdigit() else txns[i]["txn"])
        else:
            ws.cell(row_idx, 1, i + 1)
            ws.cell(row_idx, 2, None)
            ws.cell(row_idx, 3, None)

    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--out",
        default=r"C:\Users\DELL\Desktop\50 eSign Transaction Format.xlsx",
        help="Output Excel path",
    )
    parser.add_argument("--limit", type=int, default=50)
    args = parser.parse_args()

    service_root = Path(__file__).resolve().parent.parent
    mongo_uri = load_env_mongo_uri(service_root)
    txns = fetch_completed_txns(mongo_uri, args.limit)

    out_path = Path(args.out)
    write_excel(out_path, txns, target_rows=50)

    print(f"Excel updated: {out_path}")
    print(f"Genuine completed transactions: {len(txns)} / 50")
    for i, t in enumerate(txns, 1):
        print(f"  {i}. {t['date']}  txn={t['txn']}")
    if len(txns) < 50:
        print(
            f"\nNeed {50 - len(txns)} more: complete OTP on esignuat.vsign.in for each new envelope."
        )
        print("Bulk txnref script does NOT count — each txn needs real Aadhaar OTP + signed PDF.")


if __name__ == "__main__":
    main()
