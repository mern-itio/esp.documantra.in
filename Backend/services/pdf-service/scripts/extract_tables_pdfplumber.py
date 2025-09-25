#!/usr/bin/env python3
"""
Advanced PDF Table Extraction using pdfplumber, pandas, and openpyxl
This script provides robust table extraction from PDF files with multiple detection methods.
"""

import sys
import json
import os
import argparse
import time
import pandas as pd
import pdfplumber
from pathlib import Path
import logging
from typing import List, Dict, Any, Optional, Tuple
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

# Suppress warnings
warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class PDFTableExtractor:
    """
    Advanced PDF table extraction using pdfplumber with multiple detection strategies
    """
    
    def __init__(self, pdf_path: str, output_dir: str = None):
        self.pdf_path = pdf_path
        self.output_dir = output_dir or os.path.dirname(pdf_path)
        self.tables = []
        self.extraction_stats = {
            'total_pages': 0,
            'tables_found': 0,
            'extraction_method': '',
            'processing_time': 0,
            'success_rate': 0
        }
    
    def extract_tables(self, 
                      detection_method: str = 'auto',
                      output_format: str = 'xlsx',
                      preserve_formatting: bool = True,
                      extract_headers: bool = True,
                      merge_tables: bool = False,
                      page_range: Optional[str] = None,
                      language: str = 'eng') -> Dict[str, Any]:
        """
        Extract tables from PDF using multiple strategies
        
        Args:
            detection_method: 'auto', 'manual', 'all'
            output_format: 'xlsx', 'csv', 'xls'
            preserve_formatting: Whether to preserve original formatting
            extract_headers: Whether to treat first row as headers
            merge_tables: Whether to merge all tables into one file
            page_range: Page range to process (e.g., '1-5,10,15-20')
            language: Language for OCR (if needed)
        
        Returns:
            Dictionary with extraction results
        """
        start_time = time.time()
        
        try:
            # Parse page range
            pages_to_process = self._parse_page_range(page_range)
            
            with pdfplumber.open(self.pdf_path) as pdf:
                self.extraction_stats['total_pages'] = len(pdf.pages)
                
                # Process specified pages or all pages
                pages_to_extract = pages_to_process if pages_to_process else range(len(pdf.pages))
                
                for page_num in pages_to_extract:
                    if page_num >= len(pdf.pages):
                        continue
                        
                    page = pdf.pages[page_num]
                    logger.info(f"Processing page {page_num + 1}")
                    
                    # Extract tables using multiple methods
                    page_tables = self._extract_tables_from_page(
                        page, page_num + 1, detection_method, 
                        preserve_formatting, extract_headers
                    )
                    
                    self.tables.extend(page_tables)
                
                self.extraction_stats['tables_found'] = len(self.tables)
                self.extraction_stats['processing_time'] = time.time() - start_time
                self.extraction_stats['success_rate'] = len(self.tables) / max(len(pages_to_extract), 1) * 100
                
                # Generate output file
                output_path = self._generate_output_file(
                    output_format, merge_tables, extract_headers
                )
                
                return {
                    'success': True,
                    'tables_found': len(self.tables),
                    'output_file': output_path,
                    'stats': self.extraction_stats,
                    'tables': self.tables
                }
                
        except Exception as e:
            logger.error(f"Error extracting tables: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'tables_found': 0,
                'stats': self.extraction_stats
            }
    
    def _parse_page_range(self, page_range: Optional[str]) -> List[int]:
        """Parse page range string into list of page numbers"""
        if not page_range:
            return []
        
        pages = []
        ranges = page_range.split(',')
        
        for range_str in ranges:
            range_str = range_str.strip()
            if '-' in range_str:
                start, end = map(int, range_str.split('-'))
                pages.extend(range(start - 1, end))  # Convert to 0-based indexing
            else:
                pages.append(int(range_str) - 1)  # Convert to 0-based indexing
        
        return sorted(set(pages))
    
    def _extract_tables_from_page(self, page, page_num: int, detection_method: str, 
                                 preserve_formatting: bool, extract_headers: bool) -> List[Dict]:
        """Extract tables from a single page using multiple strategies"""
        tables = []
        
        try:
            # Method 1: pdfplumber's built-in table detection
            if detection_method in ['auto', 'all']:
                pdfplumber_tables = self._extract_with_pdfplumber(
                    page, page_num, preserve_formatting, extract_headers
                )
                tables.extend(pdfplumber_tables)
            
            # Method 2: Text-based table detection
            if detection_method in ['auto', 'all'] or len(tables) == 0:
                text_tables = self._extract_with_text_analysis(
                    page, page_num, preserve_formatting, extract_headers
                )
                tables.extend(text_tables)
            
            # Method 3: Line-based table detection
            if detection_method in ['auto', 'all'] or len(tables) == 0:
                line_tables = self._extract_with_line_analysis(
                    page, page_num, preserve_formatting, extract_headers
                )
                tables.extend(line_tables)
            
            # Remove duplicates and merge similar tables
            tables = self._deduplicate_tables(tables)
            
        except Exception as e:
            logger.error(f"Error extracting tables from page {page_num}: {str(e)}")
        
        return tables
    
    def _extract_with_pdfplumber(self, page, page_num: int, preserve_formatting: bool, 
                                extract_headers: bool) -> List[Dict]:
        """Extract tables using pdfplumber's built-in table detection"""
        tables = []
        
        try:
            # Try different table settings
            table_settings = [
                {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
                {"vertical_strategy": "lines", "horizontal_strategy": "text"},
                {"vertical_strategy": "text", "horizontal_strategy": "lines"},
                {"vertical_strategy": "text", "horizontal_strategy": "text"},
                {"vertical_strategy": "explicit", "horizontal_strategy": "explicit"},
            ]
            
            for i, settings in enumerate(table_settings):
                try:
                    page_tables = page.extract_tables(table_settings=settings)
                    
                    for j, table_data in enumerate(page_tables):
                        if table_data and len(table_data) > 0:
                            # Clean and process table data
                            cleaned_table = self._clean_table_data(table_data, preserve_formatting)
                            
                            if len(cleaned_table) > 0:
                                table_dict = {
                                    'name': f'Table_{page_num}_{i}_{j}',
                                    'page_number': page_num,
                                    'rows': cleaned_table,
                                    'columns': len(cleaned_table[0]) if cleaned_table else 0,
                                    'detection_method': f'pdfplumber_{i}',
                                    'confidence': 0.9
                                }
                                tables.append(table_dict)
                                
                except Exception as e:
                    logger.debug(f"pdfplumber method {i} failed: {str(e)}")
                    continue
                    
        except Exception as e:
            logger.error(f"pdfplumber extraction failed: {str(e)}")
        
        return tables
    
    def _extract_with_text_analysis(self, page, page_num: int, preserve_formatting: bool, 
                                   extract_headers: bool) -> List[Dict]:
        """Extract tables using text analysis and pattern recognition"""
        tables = []
        
        try:
            # Get text content
            text = page.extract_text()
            if not text:
                return tables
            
            # Split into lines and analyze structure
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            # Find potential table sections
            table_sections = self._identify_table_sections(lines)
            
            for i, section in enumerate(table_sections):
                if len(section) >= 2:  # At least 2 rows
                    # Parse table structure
                    parsed_table = self._parse_table_structure(section, preserve_formatting)
                    
                    if parsed_table and len(parsed_table) > 0:
                        table_dict = {
                            'name': f'Text_Table_{page_num}_{i}',
                            'page_number': page_num,
                            'rows': parsed_table,
                            'columns': len(parsed_table[0]) if parsed_table else 0,
                            'detection_method': 'text_analysis',
                            'confidence': 0.7
                        }
                        tables.append(table_dict)
                        
        except Exception as e:
            logger.error(f"Text analysis extraction failed: {str(e)}")
        
        return tables
    
    def _extract_with_line_analysis(self, page, page_num: int, preserve_formatting: bool, 
                                   extract_headers: bool) -> List[Dict]:
        """Extract tables using line and shape analysis"""
        tables = []
        
        try:
            # Get lines and shapes
            lines = page.lines
            rects = page.rects
            
            # Find potential table boundaries
            table_boundaries = self._find_table_boundaries(lines, rects)
            
            for i, boundary in enumerate(table_boundaries):
                # Extract content within boundary
                table_content = self._extract_content_in_boundary(page, boundary)
                
                if table_content and len(table_content) > 0:
                    table_dict = {
                        'name': f'Line_Table_{page_num}_{i}',
                        'page_number': page_num,
                        'rows': table_content,
                        'columns': len(table_content[0]) if table_content else 0,
                        'detection_method': 'line_analysis',
                        'confidence': 0.8
                    }
                    tables.append(table_dict)
                    
        except Exception as e:
            logger.error(f"Line analysis extraction failed: {str(e)}")
        
        return tables
    
    def _identify_table_sections(self, lines: List[str]) -> List[List[str]]:
        """Identify potential table sections in text"""
        table_sections = []
        current_section = []
        
        for line in lines:
            # Check if line looks like table data
            if self._is_table_like_line(line):
                current_section.append(line)
            else:
                if len(current_section) >= 2:
                    table_sections.append(current_section)
                current_section = []
        
        # Add final section if it exists
        if len(current_section) >= 2:
            table_sections.append(current_section)
        
        return table_sections
    
    def _is_table_like_line(self, line: str) -> bool:
        """Check if a line looks like table data"""
        # Check for multiple columns (tabs, multiple spaces, or separators)
        if '\t' in line:
            return True
        
        # Check for multiple spaces (potential column separators)
        if len(line.split('  ')) >= 3:  # At least 3 parts when split by double spaces
            return True
        
        # Check for common table patterns
        table_patterns = [
            r'\d+\s+\w+\s+\d+',  # Number Word Number
            r'\w+\s+\d+\s+\w+',  # Word Number Word
            r'.*\|\s*.*',         # Contains pipe separators
        ]
        
        for pattern in table_patterns:
            if re.search(pattern, line):
                return True
        
        return False
    
    def _parse_table_structure(self, lines: List[str], preserve_formatting: bool) -> List[List[str]]:
        """Parse table structure from lines"""
        table_data = []
        
        for line in lines:
            # Try different parsing methods
            cells = self._parse_table_row(line, preserve_formatting)
            if cells and len(cells) > 0:
                table_data.append(cells)
        
        return table_data
    
    def _parse_table_row(self, line: str, preserve_formatting: bool) -> List[str]:
        """Parse a single table row into cells"""
        # Try different separators
        separators = ['\t', '  ', '|', ',', ';']
        
        for sep in separators:
            if sep in line:
                cells = [cell.strip() for cell in line.split(sep) if cell.strip()]
                if len(cells) >= 2:  # At least 2 columns
                    return cells
        
        # Fallback: split by multiple spaces
        cells = [cell.strip() for cell in re.split(r'\s{2,}', line) if cell.strip()]
        return cells if len(cells) >= 2 else []
    
    def _find_table_boundaries(self, lines: List, rects: List) -> List[Dict]:
        """Find potential table boundaries using lines and rectangles"""
        boundaries = []
        
        # Look for rectangular shapes that could be tables
        for rect in rects:
            if rect.get('width', 0) > 100 and rect.get('height', 0) > 50:
                boundaries.append({
                    'x0': rect.get('x0', 0),
                    'y0': rect.get('y0', 0),
                    'x1': rect.get('x1', 0),
                    'y1': rect.get('y1', 0)
                })
        
        return boundaries
    
    def _extract_content_in_boundary(self, page, boundary: Dict) -> List[List[str]]:
        """Extract content within a boundary"""
        try:
            # Create a cropped page
            cropped = page.crop(boundary)
            
            # Extract tables from cropped area
            tables = cropped.extract_tables()
            
            if tables and len(tables) > 0:
                return self._clean_table_data(tables[0], True)
            
        except Exception as e:
            logger.debug(f"Boundary extraction failed: {str(e)}")
        
        return []
    
    def _clean_table_data(self, table_data: List[List[str]], preserve_formatting: bool) -> List[List[str]]:
        """Clean and normalize table data"""
        if not table_data:
            return []
        
        cleaned_data = []
        
        for row in table_data:
            if not row:
                continue
                
            # Clean each cell
            cleaned_row = []
            for cell in row:
                if cell is None:
                    cleaned_row.append('')
                else:
                    cell_str = str(cell).strip()
                    if preserve_formatting:
                        cleaned_row.append(cell_str)
                    else:
                        # Remove extra whitespace
                        cleaned_row.append(' '.join(cell_str.split()))
            
            # Only add rows with at least one non-empty cell
            if any(cell for cell in cleaned_row):
                cleaned_data.append(cleaned_row)
        
        return cleaned_data
    
    def _deduplicate_tables(self, tables: List[Dict]) -> List[Dict]:
        """Remove duplicate and merge similar tables"""
        if not tables:
            return []
        
        # Simple deduplication based on content similarity
        unique_tables = []
        seen_content = set()
        
        for table in tables:
            # Create a content hash
            content_hash = hash(str(table.get('rows', [])))
            
            if content_hash not in seen_content:
                seen_content.add(content_hash)
                unique_tables.append(table)
        
        return unique_tables
    
    def _generate_output_file(self, output_format: str, merge_tables: bool, 
                             extract_headers: bool) -> str:
        """Generate output file in specified format"""
        if not self.tables:
            # Create empty file with message
            return self._create_empty_output(output_format)
        
        timestamp = int(time.time())
        base_name = Path(self.pdf_path).stem
        
        if output_format.lower() == 'xlsx':
            output_path = os.path.join(self.output_dir, f"{base_name}_tables_{timestamp}.xlsx")
            self._create_excel_file(output_path, merge_tables, extract_headers)
        elif output_format.lower() == 'csv':
            output_path = os.path.join(self.output_dir, f"{base_name}_tables_{timestamp}.csv")
            self._create_csv_file(output_path, merge_tables, extract_headers)
        else:
            # Default to Excel
            output_path = os.path.join(self.output_dir, f"{base_name}_tables_{timestamp}.xlsx")
            self._create_excel_file(output_path, merge_tables, extract_headers)
        
        return output_path
    
    def _create_excel_file(self, output_path: str, merge_tables: bool, extract_headers: bool):
        """Create Excel file with extracted tables"""
        try:
            from openpyxl import Workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            if merge_tables and len(self.tables) > 1:
                # Merge all tables into one sheet
                ws = wb.create_sheet("Merged Tables")
                current_row = 1
                
                for i, table in enumerate(self.tables):
                    # Add table title
                    ws.cell(row=current_row, column=1, value=f"Table {i+1}: {table.get('name', '')}")
                    current_row += 1
                    
                    # Add table data
                    for row_data in table.get('rows', []):
                        for col_idx, cell_value in enumerate(row_data):
                            ws.cell(row=current_row, column=col_idx + 1, value=cell_value)
                        current_row += 1
                    
                    # Add spacing between tables
                    current_row += 1
            else:
                # Create separate sheets for each table
                for i, table in enumerate(self.tables):
                    sheet_name = f"Table_{i+1}"[:31]  # Excel sheet name limit
                    ws = wb.create_sheet(sheet_name)
                    
                    # Add table data
                    for row_idx, row_data in enumerate(table.get('rows', [])):
                        for col_idx, cell_value in enumerate(row_data):
                            ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
            
            wb.save(output_path)
            logger.info(f"Excel file created: {output_path}")
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}")
            raise
    
    def _create_csv_file(self, output_path: str, merge_tables: bool, extract_headers: bool):
        """Create CSV file with extracted tables"""
        try:
            import csv
            
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                if merge_tables and len(self.tables) > 1:
                    # Merge all tables
                    for i, table in enumerate(self.tables):
                        writer.writerow([f"Table {i+1}: {table.get('name', '')}"])
                        writer.writerow([])  # Empty row
                        
                        for row_data in table.get('rows', []):
                            writer.writerow(row_data)
                        
                        writer.writerow([])  # Empty row between tables
                else:
                    # Use first table or combine
                    table = self.tables[0] if self.tables else {'rows': []}
                    for row_data in table.get('rows', []):
                        writer.writerow(row_data)
            
            logger.info(f"CSV file created: {output_path}")
            
        except Exception as e:
            logger.error(f"Error creating CSV file: {str(e)}")
            raise
    
    def _create_empty_output(self, output_format: str) -> str:
        """Create empty output file with informative message"""
        timestamp = int(time.time())
        base_name = Path(self.pdf_path).stem
        
        if output_format.lower() == 'xlsx':
            output_path = os.path.join(self.output_dir, f"{base_name}_no_tables_{timestamp}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "No Tables Found"
            
            ws.cell(row=1, column=1, value="No tables detected in the PDF")
            ws.cell(row=2, column=1, value="The PDF may not contain table structures")
            ws.cell(row=3, column=1, value="Try different detection methods or check the PDF content")
            
            wb.save(output_path)
        else:
            output_path = os.path.join(self.output_dir, f"{base_name}_no_tables_{timestamp}.csv")
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("No tables detected\n")
                f.write("The PDF may not contain table structures\n")
                f.write("Try different detection methods or check the PDF content\n")
        
        return output_path


def main():
    """Main function for command line usage"""
    try:
        parser = argparse.ArgumentParser(description='Extract tables from PDF using pdfplumber')
        parser.add_argument('pdf_path', help='Path to PDF file')
        parser.add_argument('--output-dir', help='Output directory', default=None)
        parser.add_argument('--detection-method', choices=['auto', 'manual', 'all'], default='auto')
        parser.add_argument('--output-format', choices=['xlsx', 'csv', 'xls'], default='xlsx')
        parser.add_argument('--preserve-formatting', action='store_true', default=True)
        parser.add_argument('--extract-headers', action='store_true', default=True)
        parser.add_argument('--merge-tables', action='store_true', default=False)
        parser.add_argument('--page-range', help='Page range (e.g., 1-5,10,15-20)', default=None)
        parser.add_argument('--language', default='eng')
        parser.add_argument('--json', action='store_true', help='Output results as JSON')
        
        args = parser.parse_args()
        
        # Validate input file
        if not os.path.exists(args.pdf_path):
            error_result = {
                'success': False,
                'error': f'PDF file not found: {args.pdf_path}',
                'tables_found': 0,
                'output_file': None
            }
            if args.json:
                print(json.dumps(error_result))
            else:
                print(f"Error: PDF file not found: {args.pdf_path}")
            sys.exit(1)
        
        # Validate PDF file
        try:
            import pdfplumber
            with pdfplumber.open(args.pdf_path) as pdf:
                if len(pdf.pages) == 0:
                    raise ValueError("PDF has no pages")
        except Exception as e:
            error_result = {
                'success': False,
                'error': f'Invalid PDF file: {str(e)}',
                'tables_found': 0,
                'output_file': None
            }
            if args.json:
                print(json.dumps(error_result))
            else:
                print(f"Error: Invalid PDF file: {str(e)}")
            sys.exit(1)
        
        # Create extractor
        extractor = PDFTableExtractor(args.pdf_path, args.output_dir)
        
        # Extract tables
        result = extractor.extract_tables(
            detection_method=args.detection_method,
            output_format=args.output_format,
            preserve_formatting=args.preserve_formatting,
            extract_headers=args.extract_headers,
            merge_tables=args.merge_tables,
            page_range=args.page_range,
            language=args.language
        )
        
        # Output results
        if args.json:
            print(json.dumps(result, indent=2))
        else:
            if result['success']:
                print(f"✅ Successfully extracted {result['tables_found']} tables")
                print(f"📁 Output file: {result['output_file']}")
                print(f"📊 Processing time: {result['stats']['processing_time']:.2f} seconds")
            else:
                print(f"❌ Extraction failed: {result.get('error', 'Unknown error')}")
                sys.exit(1)
                
    except Exception as e:
        error_result = {
            'success': False,
            'error': f'Script execution failed: {str(e)}',
            'tables_found': 0,
            'output_file': None
        }
        if '--json' in sys.argv:
            print(json.dumps(error_result))
        else:
            print(f"❌ Script execution failed: {str(e)}")
        sys.exit(1)


if __name__ == '__main__':
    main()
