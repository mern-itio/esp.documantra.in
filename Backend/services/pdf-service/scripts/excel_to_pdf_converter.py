#!/usr/bin/env python3
import json
import os
import shutil
import subprocess
import sys
import tempfile
import math

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
except Exception:
    load_workbook = None
    Alignment = None
    get_column_letter = None
    PageMargins = None


def main():
    if len(sys.argv) < 3:
        print(json.dumps({
            "success": False,
            "error": "Usage: excel_to_pdf_converter.py <input.xlsx> <output.pdf>"
        }))
        return 1

    input_path = os.path.abspath(sys.argv[1])
    output_path = os.path.abspath(sys.argv[2])

    if not os.path.isfile(input_path):
        print(json.dumps({
            "success": False,
            "error": f"Input file not found: {input_path}"
        }))
        return 1

    # Ensure LibreOffice is available
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        print(json.dumps({
            "success": False,
            "error": "LibreOffice (soffice) not found in PATH"
        }))
        return 1

    # If it's an .xlsx and openpyxl is available, pre-process to wrap text and increase row heights
    source_for_export = input_path
    temp_adjusted_path = None
    ext = os.path.splitext(input_path)[1].lower()
    if ext == ".xlsx" and load_workbook is not None:
        try:
            wb = load_workbook(filename=input_path)
            for ws in wb.worksheets:
                # Page setup: fit to one page wide, unlimited height; add comfortable margins
                try:
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 0
                except Exception:
                    pass
                try:
                    if PageMargins is not None:
                        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3)
                except Exception:
                    pass

                # Estimate and expand column widths to reduce wrapping/overlap
                try:
                    if get_column_letter is not None:
                        for col_idx, column in enumerate(ws.columns, start=1):
                            max_len = 0
                            for cell in column:
                                try:
                                    val = cell.value
                                    if val is None:
                                        continue
                                    val = str(val)
                                    if len(val) > max_len:
                                        max_len = len(val)
                                except Exception:
                                    continue
                            # Rough width: characters * 0.9 plus padding, clamped
                            est_width = max(10, min(60, int(max_len * 0.9) + 2))
                            try:
                                letter = get_column_letter(col_idx)
                                current = ws.column_dimensions.get(letter).width if ws.column_dimensions.get(letter) else None
                                if current is None or current < est_width:
                                    ws.column_dimensions[letter].width = est_width
                            except Exception:
                                pass
                except Exception:
                    pass

                # Estimate characters per line if column widths are not defined
                # Default Excel row height baseline ~15 points
                default_row_height = ws.sheet_format.defaultRowHeight or 15
                for row in ws.iter_rows():
                    max_lines = 1
                    for cell in row:
                        value = cell.value
                        if value is None:
                            continue
                        if not isinstance(value, str):
                            value = str(value)
                        # Enable wrap text
                        try:
                            cell.alignment = Alignment(wrap_text=True,
                                                      horizontal=(cell.alignment.horizontal if cell.alignment else None),
                                                      vertical=(cell.alignment.vertical if cell.alignment else 'top'))
                        except Exception:
                            pass
                        # Roughly estimate lines based on length; assume ~60 chars per line average
                        hard_lines = value.count("\n") + 1
                        estimated_soft_lines = max(1, math.ceil(len(value) / 60))
                        total_lines = max(hard_lines, estimated_soft_lines)
                        if total_lines > max_lines:
                            max_lines = total_lines
                    # Increase row height proportional to lines (15 pt per line + padding)
                    new_height = max(default_row_height, 15 * max_lines + 6)
                    try:
                        ws.row_dimensions[row[0].row].height = new_height
                    except Exception:
                        pass
            temp_dir_xlsx = tempfile.mkdtemp(prefix="xlsx2pdf_pre_")
            temp_adjusted_path = os.path.join(temp_dir_xlsx, os.path.basename(input_path))
            wb.save(temp_adjusted_path)
            source_for_export = temp_adjusted_path
        except Exception:
            # If adjustment fails, fall back to original input
            source_for_export = input_path

    # Convert using LibreOffice headless mode; preserves layout, colors, images, and print setup
    tmp_dir = tempfile.mkdtemp(prefix="xlsx2pdf_")
    try:
        cmd = [
            soffice,
            "--headless",
            "--nologo",
            "--nolockcheck",
            "--nodefault",
            "--norestore",
            "--invisible",
            "--convert-to", "pdf",
            "--outdir", tmp_dir,
            source_for_export
        ]

        proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=300)

        # LibreOffice writes the output PDF to tmp_dir with same basename
        base_name = os.path.splitext(os.path.basename(source_for_export))[0]
        produced_pdf = os.path.join(tmp_dir, base_name + ".pdf")

        if proc.returncode != 0 or not os.path.exists(produced_pdf):
            print(json.dumps({
                "success": False,
                "error": "LibreOffice conversion failed",
                "stdout": proc.stdout[-1000:],
                "stderr": proc.stderr[-1000:]
            }))
            return 1

        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        # Move to requested output path
        shutil.move(produced_pdf, output_path)

        size = os.path.getsize(output_path) if os.path.exists(output_path) else 0
        print(json.dumps({
            "success": True,
            "output": output_path,
            "fileSize": size
        }))
        return 0
    except subprocess.TimeoutExpired:
        print(json.dumps({
            "success": False,
            "error": "LibreOffice conversion timed out"
        }))
        return 1
    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }))
        return 1
    finally:
        try:
            shutil.rmtree(tmp_dir)
        except Exception:
            pass
        if temp_adjusted_path:
            try:
                shutil.rmtree(os.path.dirname(temp_adjusted_path))
            except Exception:
                pass


if __name__ == "__main__":
    sys.exit(main())


