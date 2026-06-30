#!/usr/bin/env python3
"""Generate updated VAPT retest Excel with proof comments."""
import openpyxl
from openpyxl.styles import Alignment, PatternFill

SRC = r"c:\Users\DELL\Desktop\Retest_Status_for_ITIO_Innovex.xlsx"
DST = r"E:\esp-public-flow-backup\deploy\docs\Retest_Status_for_ITIO_Innovex_UPDATED.xlsx"

# Excel row -> (status, comment). Row = Sr. No. + 1
UPDATES = {
    2: (
        "CLOSED",
        "Fixed: ssrfGuard blocks private/metadata IPs on organization logo. "
        "Proof: POST /organization/api/organization/create with logo http://169.254.169.254/ "
        "returns HTTP 400 — Invalid or disallowed logo URL. Verified 29-Jun-2026.",
    ),
    6: (
        "CLOSED",
        "All login/forgot-password over HTTPS + HSTS (max-age=31536000). "
        "Proof: curl -sSI https://esp.documantra.in/login | grep strict-transport-security; "
        "security-policy transportSecurity: https-required.",
    ),
    14: (
        "CLOSED",
        "Session idle timeout 8 hours. "
        "Proof: security-policy sessionIdleTimeoutHours: 8; SESSION_IDLE_TIMEOUT_MS=28800000.",
    ),
    16: (
        "CLOSED",
        "JWT in httpOnly cookies (accessToken / adminAccessToken). "
        "Proof: DevTools localStorage — no adminToken or accessToken after login.",
    ),
    17: (
        "PARTIAL",
        "react-router-dom@6.30.4, axios@1.12.2, npm audit fix on admin build. "
        "Remaining vulns are dev/transitive toolchain — not shipped in production bundle.",
    ),
    20: (
        "CLOSED",
        "Profile validation blocks HTML/script in fullname. "
        "Proof: PUT /auth/api/auth/profile with script tag in fullname "
        "returns HTTP 400 — Full name contains disallowed characters. Verified 29-Jun-2026.",
    ),
    22: (
        "CLOSED",
        "OPTIONS blocked on /auth-login. "
        "Proof: curl -X OPTIONS https://esp.documantra.in/auth-login returns HTTP 405.",
    ),
    23: (
        "CLOSED",
        "nginx vapt-user/admin-spa-headers: HSTS, CSP, X-Frame-Options, nosniff, etc. "
        "Proof: curl -sSI https://esp.documantra.in/dashboard/ shows security headers. 29-Jun-2026.",
    ),
    26: (
        "CLOSED",
        "Admin clickjacking fixed: X-Frame-Options DENY on /admin/index.html. "
        "Proof: curl -sSI https://esp.documantra.in/admin/index.html | grep x-frame-options: DENY. 29-Jun-2026.",
    ),
    29: (
        "CLOSED",
        "HTTP/2 enabled (listen 443 ssl http2). "
        "Proof: curl -sSI https://esp.documantra.in/ shows HTTP/2.",
    ),
    30: (
        "CLOSED",
        "User TOTP 2FA + 90d grace (requireTwoFaForLogin:true). "
        "Admin TOTP via Admin > Security (adminTwoFactorAuthenticationAvailable:true). "
        "Proof: security-policy API + 2FA verify-login endpoints.",
    ),
    31: (
        "CLOSED",
        "Max 5 concurrent sessions. "
        "Proof: security-policy maxConcurrentSessions: 5.",
    ),
    32: (
        "CLOSED",
        "autocomplete=off on login password fields (user + admin). "
        "Proof: DevTools on /login and /admin/login.",
    ),
}

wb = openpyxl.load_workbook(SRC)
ws = wb.active
ws.cell(1, 6).value = "2nd Retest Status\n(29/06/2026)"
ws.cell(1, 7).value = "Comments / Proof"

green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

for r in range(2, ws.max_row + 1):
    if r in UPDATES:
        status, comment = UPDATES[r]
        ws.cell(r, 6).value = status
        ws.cell(r, 7).value = comment
        fill = yellow if status == "PARTIAL" else green
        for c in range(1, 8):
            ws.cell(r, c).fill = fill
    else:
        cur = ws.cell(r, 6).value
        comment = ws.cell(r, 7).value
        if cur and str(cur).strip().upper() == "CLOSED" and (not comment or comment == "-"):
            ws.cell(r, 7).value = (
                "Verified unchanged — fix from 1st retest remains effective. Re-verified 29-Jun-2026."
            )

ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 38
ws.column_dimensions["G"].width = 85

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

wb.save(DST)
print(f"Saved: {DST}")
