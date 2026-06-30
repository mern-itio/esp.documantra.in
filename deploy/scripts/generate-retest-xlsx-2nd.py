#!/usr/bin/env python3
"""Fill 2nd retest Excel (30-Jun-2026) with evidence comments and updated status."""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SRC = r"c:\Users\DELL\Desktop\Retest_Status_for_ITIO_Innovex-new.xlsx"
DST_DESKTOP = r"c:\Users\DELL\Desktop\Retest_Status_for_ITIO_Innovex-new_WITH_EVIDENCE.xlsx"
DST_REPO = r"E:\esp-public-flow-backup\deploy\docs\Retest_Status_for_ITIO_Innovex-new_WITH_EVIDENCE.xlsx"

# Excel row -> (2nd_retest_status, evidence_comment)
# Row = Sr. No. + 1
UPDATES = {
    2: (
        "CLOSED",
        "Fixed: ssrfGuard + resolveLogoInput on organization logo URL. "
        "Proof (30-Jun-2026): POST /organization/api/organization/create with logo "
        "http://169.254.169.254/ returns HTTP 400 — Invalid or disallowed logo URL.",
    ),
    3: (
        "CLOSED",
        "Password change invalidates all active sessions (activeSessions cleared, passwordChangedAt updated). "
        "Verified unchanged from 1st retest — re-checked 30-Jun-2026 on production.",
    ),
    4: (
        "CLOSED",
        "Email change invalidates existing sessions (emailChangedAt check on session validation). "
        "Verified unchanged — re-checked 30-Jun-2026.",
    ),
    5: (
        "CLOSED",
        "Forgot/reset password enforces min 8 chars + upper/lower/number/special. "
        "Verified unchanged — re-checked 30-Jun-2026.",
    ),
    6: (
        "CLOSED",
        "All credential traffic over HTTPS only. Proof (30-Jun-2026 server curl): "
        "curl -sI http://esp.documantra.in/login → HTTP 301 Moved Permanently to HTTPS; "
        "curl -sI https://esp.documantra.in/login → strict-transport-security: max-age=31536000; includeSubDomains; "
        "GET /auth/api/auth/security-policy → transportSecurity: https-required. "
        "Login POST only on https:// — no cleartext transmission.",
    ),
    7: (
        "CLOSED",
        "Profile fullname validation blocks HTML/script. "
        "Proof: PUT /auth/api/auth/profile with <script> in fullname → HTTP 400. Verified 30-Jun-2026.",
    ),
    8: (
        "CLOSED",
        "Admin subscription credits validated server-side (negative values rejected). "
        "Verified unchanged — re-checked 30-Jun-2026.",
    ),
    9: (
        "CLOSED",
        "Price tampering blocked by server-side validation on subscription APIs. "
        "Verified unchanged — re-checked 30-Jun-2026.",
    ),
    10: (
        "CLOSED",
        "Rate limiting on admin add-payment endpoint. Verified unchanged — re-checked 30-Jun-2026.",
    ),
    11: (
        "CLOSED",
        "Rate limiting on admin new-template endpoint. Verified unchanged — re-checked 30-Jun-2026.",
    ),
    12: (
        "CLOSED",
        "Rate limiting on admin add-PDF-tool endpoint. Verified unchanged — re-checked 30-Jun-2026.",
    ),
    13: (
        "CLOSED",
        "Password history policy on admin forgot-password flow. Verified unchanged — re-checked 30-Jun-2026.",
    ),
    14: (
        "CLOSED",
        "Session idle timeout enforced server-side and configurable from Admin → Sessions. "
        "Proof: GET /auth/api/auth/security-policy → sessionIdleTimeoutHours; "
        "Admin dashboard Sessions page saves policy (idle timeout hours + max concurrent sessions). "
        "Idle sessions invalidated via lastActive check on validateSession.",
    ),
    15: (
        "CLOSED",
        "JWT/session tokens stored in httpOnly Secure cookies (M14) — not in localStorage. "
        "Proof: DevTools → Application → Local Storage after login — no accessToken/JWT key. "
        "Auth uses credentials:include + adminAccessToken cookie.",
    ),
    16: (
        "CLOSED",
        "Fixed 30-Jun-2026: userProfileSnapshot removed from sessionStorage. "
        "Profile hint stored in-memory only (user id); JWT in httpOnly cookie. "
        "Proof: DevTools → Application → Session Storage after login — no userProfileSnapshot key. "
        "Legacy keys purged on load (userProfileSnapshot, userData, accessToken).",
    ),
    17: (
        "OPEN",
        "Deferred — nodemailer upgrade planned for a later release. "
        "Current state: auth-service npm audit may report nodemailer HIGH until upgraded.",
    ),
    18: (
        "CLOSED",
        "Generic error messages on login/forgot-password (no user enumeration). "
        "Verified unchanged — re-checked 30-Jun-2026.",
    ),
    19: (
        "CLOSED",
        "Input validation on e-sign create template. Verified unchanged — re-checked 30-Jun-2026.",
    ),
    20: (
        "CLOSED",
        "Profile field validation (getPlainTextFieldError) blocks disallowed characters. "
        "Proof: PUT /auth/api/auth/profile with script tag → HTTP 400. Verified 30-Jun-2026.",
    ),
    21: (
        "CLOSED",
        "AI template stream errors sanitized (no stack traces to client). "
        "Verified unchanged — re-checked 30-Jun-2026.",
    ),
    22: (
        "CLOSED",
        "OPTIONS blocked on sensitive routes. Proof (30-Jun-2026 server curl): "
        "curl -sI -X OPTIONS https://esp.documantra.in/auth-login → HTTP/2 405; "
        "curl -sI -X OPTIONS https://esp.documantra.in/login → HTTP/2 405; "
        "curl -sI -X OPTIONS https://esp.documantra.in/dashboard/ → HTTP/2 405; "
        "curl -sI -X OPTIONS https://esp.documantra.in/ → HTTP/2 405 (nginx location = / fix).",
    ),
    23: (
        "CLOSED",
        "Security headers on user/admin SPA. Proof: curl -sSI https://esp.documantra.in/dashboard/ "
        "shows HSTS, CSP, X-Frame-Options, X-Content-Type-Options. "
        "Admin /admin/index.html → X-Frame-Options: DENY. Verified 30-Jun-2026.",
    ),
    24: (
        "CLOSED",
        "nginx server_tokens off; proxy_hide_header Server on API routes. "
        "Verified unchanged — re-checked 30-Jun-2026.",
    ),
    25: (
        "CLOSED",
        "Custom error pages — no server version in OTP error responses. "
        "Verified unchanged — re-checked 30-Jun-2026.",
    ),
    26: (
        "CLOSED",
        "Admin clickjacking mitigated. Proof: curl -sSI https://esp.documantra.in/admin/index.html "
        "→ X-Frame-Options: DENY. Verified 30-Jun-2026.",
    ),
    27: (
        "CLOSED",
        "Cross-Origin-Opener-Policy: same-origin-allow-popups on nginx. "
        "Verified unchanged — re-checked 30-Jun-2026.",
    ),
    28: (
        "CLOSED",
        "Custom 404.html served for unknown routes. Verified unchanged — re-checked 30-Jun-2026.",
    ),
    29: (
        "CLOSED",
        "HTTP/2 enabled on nginx (listen 443 ssl http2). "
        "Proof: curl -sSI https://esp.documantra.in/ shows HTTP/2. Verified 30-Jun-2026.",
    ),
    30: (
        "CLOSED",
        "2FA (TOTP) enabled for users. Proof (30-Jun-2026): "
        "GET /auth/api/auth/security-policy → requireTwoFaForLogin: true, "
        "twoFactorAuthenticationAvailable: true, requireTwoFaGraceDays: 90. "
        "auth-service .env REQUIRE_2FA_FOR_LOGIN=true. "
        "Login flow returns TWO_FA_REQUIRED; test user pawneshk@itio.in requires authenticator OTP. "
        "Admin TOTP available under Admin → Security (adminTwoFactorAuthenticationAvailable: true).",
    ),
    31: (
        "CLOSED",
        "Concurrent session limit enforced. Proof (30-Jun-2026): "
        "security-policy maxConcurrentSessions: 5; enforceConcurrentSessionLimit() on login "
        "drops oldest session when limit exceeded.",
    ),
    32: (
        "CLOSED",
        "autocomplete=off on login form and password fields (user + admin). "
        "Proof: DevTools Elements on /login — form autoComplete=off, password autoComplete=off. "
        "Verified 30-Jun-2026.",
    ),
    33: (
        "CLOSED",
        "Backend Docker ports bound to 127.0.0.1; only 22/80/443 public. "
        "Proof: curl http://157.230.231.148:2101 → connection refused. "
        "DigitalOcean firewall + UFW. Verified 30-Jun-2026.",
    ),
}

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def fill_status_color(status: str):
    s = (status or "").strip().upper()
    if s == "OPEN":
        return RED
    if s == "PARTIAL":
        return YELLOW
    return GREEN


def process(src: str, dst: str) -> None:
    wb = openpyxl.load_workbook(src)
    ws = wb.active

    ws.cell(1, 7).value = "2nd Retest Status\n(30/06/2026 — ITIO verified)"
    ws.cell(1, 8).value = "Comments / Proof (30-Jun-2026)"

    for r in range(1, ws.max_row + 1):
        if r == 1:
            for c in range(1, 9):
                ws.cell(r, c).fill = HEADER
                ws.cell(r, c).font = Font(bold=True)
            continue

        if r in UPDATES:
            status, comment = UPDATES[r]
            ws.cell(r, 7).value = status
            ws.cell(r, 8).value = comment
            fill = fill_status_color(status)
            for c in range(1, 9):
                ws.cell(r, c).fill = fill

    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 95

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(dst)
    print(f"Saved: {dst}")


if __name__ == "__main__":
    process(SRC, DST_DESKTOP)
    process(SRC, DST_REPO)
    open_count = sum(1 for s, _ in UPDATES.values() if s == "OPEN")
    partial_count = sum(1 for s, _ in UPDATES.values() if s == "PARTIAL")
    closed_count = sum(1 for s, _ in UPDATES.values() if s == "CLOSED")
    print(f"Summary: {closed_count} CLOSED, {partial_count} PARTIAL, {open_count} OPEN")
